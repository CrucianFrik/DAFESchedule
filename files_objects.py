import string
from openpyxl import load_workbook
import pandas as pd


def find_not_NaN(start, end, sheet):
    for c in range(start[0], end[0] + 1):
        for r in range(start[1], end[1] + 1):
            if type(sheet.cell(row=r, column=c).value) == str and sheet.cell(row=r, column=c).value:
                return sheet.cell(row=r, column=c).value


class XlsxFile:
    def __init__(self, path):
        self.__LATIN_ALPHABET = ['-'] + list(string.ascii_uppercase)

        self.__path = path
        self.__wb = load_workbook(self.__path)

    def unmerge_cells(self):
        print("XlsxFile: unmerging started")
        for sheet_name in self.__wb.get_sheet_names():
            sheet = self.__wb.get_sheet_by_name(sheet_name)
            self.__unmerge_cells_in_list(sheet, sheet_name)
        self.save()
        print("XlsxFile: completed")

    def save(self, path=''):
        self.__wb.save(path if path else self.__path)

    def get_sheet(self, name):
        try:
            return pd.read_excel(self.__path, sheet_name=name)
        except ValueError as e:
            print(f"\nWRRNING!\n no list '{name}', lists names:", self.__wb.get_sheet_names(), "\n")

    def get_pd(self):
        sheet_list = []
        for sheet_name in self.__wb.get_sheet_names():
            sheet_list.append(pd.read_excel(self.__path, sheet_name=sheet_name))
        return sheet_list

    def __format_range_name(self, _range):  # "C7:H7" -> ((3, 7), (8, 7))
        c = 0
        r = ''
        for i in _range:
            if i.isalpha():
                c += self.__LATIN_ALPHABET.index(i)
            else:
                r += i
        return c, int(r)

    def __unmerge_cells_in_list(self, sheet, sheet_name):
        merged_cells = list(sheet.merged_cells.ranges)
        for _range in merged_cells:
            start, end = map(self.__format_range_name, str(_range).split(':'))
            sheet.unmerge_cells(str(_range))
            v = find_not_NaN(start, end, sheet)

            for c in range(start[0], end[0] + 1):
                for r in range(start[1], end[1] + 1):
                    try:
                        sheet.cell(row=r, column=c).value = sheet.cell(row=start[1], column=start[0]).value
                    except Exception as e:
                        print(f"unmerge warning in sheet: {sheet_name}, range: {_range}")
